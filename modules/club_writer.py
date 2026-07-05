import shutil
from datetime import datetime

from openpyxl import load_workbook

from modules.constants import SHEET_CLUB


class ClubWriter:

    def _backup(self, excel_datei):

        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_club_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        shutil.copy2(excel_datei, backup)

        return backup

    def _headers(self, ws):

        return {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value
        }

    def save_club(self, excel_datei, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)

        ws = wb[SHEET_CLUB]

        headers = self._headers(ws)

        if ws.max_row == 1:
            row = 2
        else:
            row = 2

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=row,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)        