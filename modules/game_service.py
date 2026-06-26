"""
=========================================================
ASV Manager
Game Service
=========================================================
Zentrale Verwaltung aller Spiele.
"""

from openpyxl import load_workbook

from modules.excel_reader import ExcelReader
from modules.excel_writer import ExcelWriter
from modules.history_service import HistoryService


class GameService:

    def __init__(self):
        self.reader = ExcelReader()
        self.writer = ExcelWriter()
        self.history = HistoryService()

    def load_games(self, excel_datei):
        self.ensure_game_ids(excel_datei)
        self.reader.load(excel_datei)
        return self.reader.df.copy()

    def update_game(self, excel_datei, excel_zeile, daten):
        game_id = self.get_game_id_by_excel_row(
            excel_datei,
            excel_zeile
        )

        self.writer.update_game(
            excel_datei,
            excel_zeile,
            daten
        )

        objekt = (
            f"{daten.get('LIGA', '')} | "
            f"{daten.get('DATUM', '')} | "
            f"{daten.get('GEGNER', '')}"
        )

        self.history.log(
            excel_datei=excel_datei,
            bereich="Spiel",
            aktion="Bearbeitet",
            objekt=objekt,
            excel_zeile=excel_zeile,
            game_id=game_id,
            grund="",
            bemerkung="Spiel wurde bearbeitet",
            benutzer="System"
        )

    def add_game(self, excel_datei, daten):
        self.writer.add_game(
            excel_datei,
            daten
        )

    def ensure_game_ids(self, excel_datei):
        wb = load_workbook(excel_datei)
        ws = wb["ICS2"]

        header = [cell.value for cell in ws[1]]

        if "GAME_ID" not in header:
            ws.insert_cols(1)
            ws["A1"] = "GAME_ID"

        game_id_col = 1
        laufende_nummer = 1

        for row in range(2, ws.max_row + 1):
            zelle = ws.cell(row=row, column=game_id_col)

            if not zelle.value:
                zelle.value = f"GAME{laufende_nummer:06d}"

            laufende_nummer += 1

        wb.save(excel_datei)

    def get_game_id_by_excel_row(self, excel_datei, excel_zeile):
        wb = load_workbook(excel_datei, data_only=True)
        ws = wb["ICS2"]

        headers = [cell.value for cell in ws[1]]

        if "GAME_ID" not in headers:
            return ""

        game_id_col = headers.index("GAME_ID") + 1

        return ws.cell(
            row=excel_zeile,
            column=game_id_col
        ).value or ""

        # -------------------------------------------------

    def archive_game(
        self,
        excel_datei,
        excel_zeile,
        grund,
        bemerkung
    ):

        game_id = self.get_game_id_by_excel_row(
            excel_datei,
            excel_zeile
        )

        self.writer.update_fields(
            excel_datei,
            excel_zeile,
            {
                "STATUS": "Archiviert"
            }
        )

        self.history.log(
            excel_datei=excel_datei,
            bereich="Spiel",
            aktion="Archiviert",
            objekt=f"Spiel {game_id}",
            excel_zeile=excel_zeile,
            game_id=game_id,
            grund=grund,
            bemerkung=bemerkung,
            benutzer="System"
        )    