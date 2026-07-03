from openpyxl import load_workbook

from modules.constants import SHEET_TRAINING_PARTICIPANTS


class TrainingParticipantWriter:

    def deactivate_participants(
        self,
        excel_datei,
        training_id
    ):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TRAINING_PARTICIPANTS]

        headers = {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
        }

        for row in range(2, ws.max_row + 1):

            if str(ws.cell(
                row=row,
                column=headers["TRAINING_ID"]
            ).value) == str(training_id):

                ws.cell(
                    row=row,
                    column=headers["AKTIV"]
                ).value = "Nein"

        wb.save(excel_datei)

    def add_participant(
        self,
        excel_datei,
        training_id,
        member_id,
        role
    ):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TRAINING_PARTICIPANTS]

        ws.append([
            training_id,
            member_id,
            role,
            "",
            "",
            "Ja"
        ])

        wb.save(excel_datei)