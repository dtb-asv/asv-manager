from openpyxl import load_workbook

from modules.constants import SHEET_TEAM_ASSIGNMENTS


class TeamAssignmentWriter:

    def deactivate_assignments(self, excel_datei, team_id):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TEAM_ASSIGNMENTS]

        headers = [cell.value for cell in ws[1]]

        team_col = headers.index("TEAM_ID") + 1
        aktiv_col = headers.index("AKTIV") + 1

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=team_col).value == team_id:
                ws.cell(row=row, column=aktiv_col).value = "NEIN"

        wb.save(excel_datei)

    def add_assignment(self, excel_datei, team_id, member_id, role_code):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TEAM_ASSIGNMENTS]

        ws.append([
            team_id,
            member_id,
            role_code,
            "",
            "",
            "JA",
            ""
        ])

        wb.save(excel_datei)