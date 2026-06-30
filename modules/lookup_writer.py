from openpyxl import load_workbook

from modules.constants import SHEET_CFG_LOOKUPS


class LookupWriter:

    def add_lookup(self, excel_datei, daten):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_CFG_LOOKUPS]

        ws.append([
            daten["LOOKUP_ID"],
            daten["LOOKUP_TYPE"],
            daten["CODE"],
            daten["NAME"],
            daten["BESCHREIBUNG"],
            daten["SORTIERUNG"],
            daten["AKTIV"],
            daten["SYSTEM"]
        ])

        wb.save(excel_datei)

    def update_lookup(self, excel_datei, lookup_id, daten):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_CFG_LOOKUPS]

        headers = [cell.value for cell in ws[1]]
        id_col = headers.index("LOOKUP_ID") + 1

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=id_col).value == lookup_id:

                for col, header in enumerate(headers, start=1):

                    if header in daten:
                        ws.cell(row=row, column=col).value = daten[header]

                break

        wb.save(excel_datei)

    def archive_lookup(self, excel_datei, lookup_id):

        self.update_lookup(
            excel_datei,
            lookup_id,
            {
                "AKTIV": "NEIN"
            }
        )