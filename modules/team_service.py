from openpyxl import load_workbook
from modules.team_writer import TeamWriter
from modules.history_service import HistoryService
from modules.department_service import DepartmentService
from modules.repositories.team_repository import TeamRepository
import pandas as pd

from modules.constants import (
    SHEET_TEAMS,
    SHEET_TEAM_VEREINE,
    SHEET_TEAM_TRAINER
)


class TeamService:

    def __init__(self):

        self.repository = TeamRepository()
        self.writer = TeamWriter() 
        self.history = HistoryService()

    def ensure_sheets(self, excel_datei):

        wb = load_workbook(excel_datei)

        if SHEET_TEAMS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_TEAMS)
            ws.append([
                "TEAM_ID",
                "NAME",
                "ALTERSKLASSE",
                "DEPARTMENT_ID",
                "AKTIV",
                "BEMERKUNG"
            ])

        if SHEET_TEAM_VEREINE not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_TEAM_VEREINE)
            ws.append([
                "TEAM_ID",
                "VEREIN",
                "ROLLE"
            ])

        if SHEET_TEAM_TRAINER not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_TEAM_TRAINER)
            ws.append([
                "TEAM_ID",
                "MEMBER_ID",
                "ROLLE"
            ])

        wb.save(excel_datei)

       

    def load_teams(self, excel_datei):

        self.ensure_sheets(excel_datei)

        return pd.read_excel(
            excel_datei,
            sheet_name=SHEET_TEAMS
        )

    def next_team_id(self, excel_datei):

        df = self.load_teams(excel_datei)

        if df.empty or "TEAM_ID" not in df.columns:
            return "TEAM000001"

        nummern = []

        for value in df["TEAM_ID"].dropna():

            text = str(value)

            if text.startswith("TEAM"):

                try:
                    nummern.append(
                        int(text.replace("TEAM", ""))
                    )
                except ValueError:
                    pass

        next_number = max(nummern, default=0) + 1

        return f"TEAM{next_number:06d}"

    def add_team(self, excel_datei, daten):

        daten["TEAM_ID"] = self.next_team_id(excel_datei)
        daten["AKTIV"] = "Ja"

        self.writer.add_team(
            excel_datei,
            daten
        )   

    def update_team(self, excel_datei, team_id, daten):

        grund = daten.pop("_GRUND", "")
        bemerkung = daten.pop("_BEMERKUNG", "")

        self.writer.update_team(
            excel_datei,
            team_id,
            daten
        )

        objekt = daten.get("NAME", team_id)

        self.history.log(
            excel_datei=excel_datei,
            bereich="Mannschaft",
            aktion="Bearbeitet",
            objekt=objekt,
            excel_zeile="",
            game_id=team_id,
            grund=grund,
            bemerkung=bemerkung,
            benutzer="System"
        )

    def archive_team(self, excel_datei, team_id):

        self.writer.archive_team(
            excel_datei,
            team_id
        )

        self.history.log(
            excel_datei=excel_datei,
            bereich="Mannschaft",
            aktion="Archiviert",
            objekt=team_id,
            excel_zeile="",
            game_id=team_id,
            grund="",
            bemerkung="Mannschaft wurde archiviert",
            benutzer="System"
        )

    def load_teams_with_department_name(self, excel_datei):

        teams = self.load_teams(excel_datei)
        departments = DepartmentService().load_departments(excel_datei)

        if teams.empty:
            return teams

        return teams.merge(
            departments[["DEPARTMENT_ID", "NAME"]],
            on="DEPARTMENT_ID",
            how="left",
            suffixes=("", "_DEPARTMENT")
        )  

    def get_active(self):
        return self.repository.get_active()    

    def get_all(self):
        return self.repository.get_all()


    def count(self):
        return self.repository.count()


    def create_team(self, name, season_id, active=True):
        return self.repository.save(
            name=name,
            season_id=season_id,
            active=active
        )


    def update_team_db(self, team_id, name, season_id, active=True):
        return self.repository.update(
            team_id=team_id,
            name=name,
            season_id=season_id,
            active=active
        )


    def archive_team_db(self, team_id):
        return self.repository.archive(team_id)      

    