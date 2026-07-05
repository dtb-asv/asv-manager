import shutil
from datetime import datetime
from openpyxl import load_workbook

from modules.constants import SHEET_TRAINING_SCHEDULES


class TrainingScheduleWriter:

    def _backup(self, excel_datei):

        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_training_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        shutil.copy2(excel_datei, backup)

        return backup

    def _headers(self, ws):

        return {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value
        }

    def _find_row_by_schedule_id(self, ws, schedule_id):

        headers = self._headers(ws)
        id_col = headers.get("SCHEDULE_ID")

        if not id_col:
            return None

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=id_col).value == schedule_id:
                return row

        return None

    def add_schedule(self, excel_datei, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TRAINING_SCHEDULES]

        headers = self._headers(ws)
        neue_zeile = ws.max_row + 1

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=neue_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return daten["SCHEDULE_ID"]

    def update_schedule(self, excel_datei, schedule_id, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_TRAINING_SCHEDULES]

        headers = self._headers(ws)
        zeile = self._find_row_by_schedule_id(ws, schedule_id)

        if zeile is None:
            raise ValueError(f"Trainingsplan {schedule_id} nicht gefunden.")

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

    def archive_schedule(self, excel_datei, schedule_id):

        self.update_schedule(
            excel_datei,
            schedule_id,
            {
                "AKTIV": "Nein"
            }
        )