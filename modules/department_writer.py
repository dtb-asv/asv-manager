import shutil
from datetime import datetime
from openpyxl import load_workbook

from modules.constants import SHEET_DEPARTMENTS


class DepartmentWriter:

    def _backup(self, excel_datei):

        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_department_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        shutil.copy2(excel_datei, backup)

        return backup

    def _headers(self, ws):

        return {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value
        }

    def _find_row_by_department_id(self, ws, department_id):

        headers = self._headers(ws)
        id_col = headers.get("DEPARTMENT_ID")

        if not id_col:
            return None

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=id_col).value == department_id:
                return row

        return None

    def add_department(self, excel_datei, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_DEPARTMENTS]

        headers = self._headers(ws)
        neue_zeile = ws.max_row + 1

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=neue_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return daten["DEPARTMENT_ID"]

    def update_department(self, excel_datei, department_id, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_DEPARTMENTS]

        headers = self._headers(ws)
        zeile = self._find_row_by_department_id(ws, department_id)

        if zeile is None:
            raise ValueError(f"Bereich {department_id} nicht gefunden.")

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

    def archive_department(self, excel_datei, department_id):

        self.update_department(
            excel_datei,
            department_id,
            {
                "AKTIV": "Nein"
            }
        )