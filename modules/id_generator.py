from openpyxl import load_workbook


class IdGenerator:

    @staticmethod
    def next_id(
        excel_datei,
        sheet_name,
        id_column,
        prefix
    ):

        wb = load_workbook(excel_datei, read_only=True)
        ws = wb[sheet_name]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        if id_column not in headers:
            wb.close()
            raise ValueError(
                f"Spalte '{id_column}' nicht gefunden."
            )

        id_col = headers.index(id_column) + 1

        nummern = []

        for row in range(2, ws.max_row + 1):

            value = ws.cell(
                row=row,
                column=id_col
            ).value

            if not value:
                continue

            text = str(value)

            if text.startswith(prefix):

                try:
                    nummern.append(
                        int(
                            text.replace(prefix, "")
                        )
                    )
                except ValueError:
                    pass

        wb.close()

        next_number = max(
            nummern,
            default=0
        ) + 1

        return f"{prefix}{next_number:06d}"