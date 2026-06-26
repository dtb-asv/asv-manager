"""
=========================================================
ASV Manager
History Service
=========================================================
"""

from datetime import datetime
from openpyxl import load_workbook


class HistoryService:

    SHEET_NAME = "HISTORY"

    def log(
        self,
        excel_datei,
        bereich,
        aktion,
        objekt,
        excel_zeile,
        game_id="",
        grund="",
        bemerkung="",
        benutzer="System"
    ):

        wb = load_workbook(excel_datei)

        if self.SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(self.SHEET_NAME)

            ws.append([
                "Zeitpunkt",
                "GAME_ID",
                "Bereich",
                "Aktion",
                "Objekt",
                "ExcelZeile",
                "Grund",
                "Bemerkung",
                "Benutzer"
            ])

        else:
            ws = wb[self.SHEET_NAME]

            headers = [cell.value for cell in ws[1]]

            if "GAME_ID" not in headers:
                ws.insert_cols(2)
                ws["B1"] = "GAME_ID"

        ws.append([
            datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            game_id,
            bereich,
            aktion,
            objekt,
            excel_zeile,
            grund,
            bemerkung,
            benutzer
        ])

        wb.save(excel_datei)