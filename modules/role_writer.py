from openpyxl import load_workbook

from modules.constants import SHEET_CFG_ROLES


class RoleWriter:

    def add_role(self, excel_datei, daten):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_CFG_ROLES]

        ws.append([
            daten["ROLE_ID"],
            daten["NAME"],
            daten["BESCHREIBUNG"],
            daten["AKTIV"]
        ])

        wb.save(excel_datei)

    def update_role(self, excel_datei, role_id, daten):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_CFG_ROLES]

        headers = [cell.value for cell in ws[1]]

        role_col = headers.index("ROLE_ID") + 1

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=role_col).value == role_id:

                for col, header in enumerate(headers, start=1):

                    if header in daten:
                        ws.cell(
                            row=row,
                            column=col
                        ).value = daten[header]

                break

        wb.save(excel_datei)

    def archive_role(self, excel_datei, role_id):

        self.update_role(
            excel_datei,
            role_id,
            {
                "AKTIV": "NEIN"
            }
        )