"""
=========================================================
ASV Manager
Excel Writer
=========================================================
"""

import shutil
from datetime import datetime
from openpyxl import load_workbook


class ExcelWriter:

    def __init__(self, sheet_name="ICS2"):
        self.sheet_name = sheet_name

    # -------------------------------------------------

    def _backup(self, excel_datei):

        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        shutil.copy2(excel_datei, backup)

        return backup

    # -------------------------------------------------

    def _headers(self, ws):

        headers = {}

        for col in range(1, ws.max_column + 1):

            name = ws.cell(row=1, column=col).value

            if name:
                headers[name] = col

        return headers

    # -------------------------------------------------

    def add_game(self, excel_datei, daten):

        backup = self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[self.sheet_name]

        headers = self._headers(ws)

        neue_zeile = ws.max_row + 1

        for feld, wert in daten.items():

            if feld in headers:

                ws.cell(
                    row=neue_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return backup

    # -------------------------------------------------

    def update_game(
        self,
        excel_datei,
        excel_zeile,
        daten
    ):

        backup = self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[self.sheet_name]

        headers = self._headers(ws)

        for feld, wert in daten.items():

            if feld in headers:

                ws.cell(
                    row=excel_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return backup

    # -------------------------------------------------

    def update_fields(
        self,
        excel_datei,
        excel_zeile,
        felder
    ):

        backup = self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[self.sheet_name]

        headers = self._headers(ws)

        for feld, wert in felder.items():

            if feld in headers:

                ws.cell(
                    row=excel_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return backup    

    # -------------------------------------------------