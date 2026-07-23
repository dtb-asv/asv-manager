"""
=========================================================
ASV Manager
Dashboard Service
=========================================================

Bereitet die Daten für die Vereinszentrale auf.
"""

from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from modules.team_service import TeamService
from modules.facility_service import FacilityService
from modules.place_service import PlaceService

from modules.constants import (
    SHEET_GAMES,
    SHEET_MEMBERS,
    SHEET_TEAMS,
    SHEET_TRAININGS,
    SHEET_TRAINING_SCHEDULES,
)


class DashboardService:

    def load(self, excel_datei):
        """
        Liefert die vollständige Übersicht für das Dashboard.
        """

        heute = date.today()
        wochenbeginn = heute - timedelta(days=heute.weekday())
        wochenende = wochenbeginn + timedelta(days=6)

        spiele = self._read_sheet(excel_datei, SHEET_GAMES)
        trainings = self._read_sheet(excel_datei, SHEET_TRAININGS)
        trainingsplaene = self._read_sheet(
            excel_datei,
            SHEET_TRAINING_SCHEDULES
        )
        mitglieder = self._read_sheet(excel_datei, SHEET_MEMBERS)
        mannschaften = self._read_sheet(excel_datei, SHEET_TEAMS)

        spiele_heute = self._filter_date(
            spiele,
            "DATUM",
            heute,
            heute
        )

        spiele_woche = self._filter_date(
            spiele,
            "DATUM",
            wochenbeginn,
            wochenende
        )

        trainings_heute = self._filter_date(
            trainings,
            "DATUM",
            heute,
            heute
        )

        trainings_woche = self._filter_date(
            trainings,
            "DATUM",
            wochenbeginn,
            wochenende
        )

        naechstes_spiel = self._next_event(
            spiele,
            date_column="DATUM",
            time_column="STARTZEIT"
        )

        naechstes_training = self._next_event(
            trainings,
            date_column="DATUM",
            time_column="STARTZEIT"
        )

        aktive_trainingsplaene = self._count_active(
            trainingsplaene
        )

        aktive_mitglieder = self._count_active(
            mitglieder
        )

        aktive_mannschaften = self._count_active(
            mannschaften
        )

        warnungen = self._create_alerts(
            spiele=spiele,
            trainings=trainings,
            trainingsplaene=trainingsplaene,
            mitglieder=mitglieder,
            mannschaften=mannschaften
        )

        return {
            "today": {
                "date": heute,
                "games": self._records(spiele_heute),
                "trainings": self._records(trainings_heute),
                "game_count": len(spiele_heute),
                "training_count": len(trainings_heute),
            },

            "week": {
                "start": wochenbeginn,
                "end": wochenende,
                "game_count": len(spiele_woche),
                "training_count": len(trainings_woche),
                "games": self._records(spiele_woche),
                "trainings": self._records(trainings_woche),
            },

            "next": {
                "game": naechstes_spiel,
                "training": naechstes_training,
            },

            "training": {
                "active_schedules": aktive_trainingsplaene,
            },

            "quick_stats": {
                "members": aktive_mitglieder,
                "teams": aktive_mannschaften,
                "games_this_week": len(spiele_woche),
                "trainings_this_week": len(trainings_woche),
                "trainings_today": len(trainings_heute),
            },

            "alerts": warnungen,

            "status": {
                "level": self._status_level(warnungen),
                "alert_count": len(warnungen),
            }
        }

    def _read_sheet(self, excel_datei, sheet_name):
        """
        Liest ein Excel-Blatt ein.

        Existiert das Blatt nicht, wird ein leerer DataFrame
        zurückgegeben. Das Dashboard bleibt dadurch startfähig.
        """

        try:
            wb = load_workbook(
                excel_datei,
                read_only=True,
                data_only=True
            )

            sheet_exists = sheet_name in wb.sheetnames
            wb.close()

            if not sheet_exists:
                return pd.DataFrame()

            return pd.read_excel(
                excel_datei,
                sheet_name=sheet_name
            )

        except Exception:
            return pd.DataFrame()

    def _filter_date(
        self,
        dataframe,
        date_column,
        start_date,
        end_date
    ):
        """
        Filtert einen DataFrame nach einem Datumsbereich.
        """

        if dataframe.empty:
            return dataframe.copy()

        if date_column not in dataframe.columns:
            return dataframe.iloc[0:0].copy()

        result = dataframe.copy()

        result["_DASHBOARD_DATE"] = result[date_column].apply(
            self._to_date
        )

        result = result[
            result["_DASHBOARD_DATE"].notna()
        ]

        result = result[
            (result["_DASHBOARD_DATE"] >= start_date)
            & (result["_DASHBOARD_DATE"] <= end_date)
        ]

        return result.drop(
            columns=["_DASHBOARD_DATE"],
            errors="ignore"
        )

    def _next_event(
        self,
        dataframe,
        date_column,
        time_column
    ):
        """
        Sucht das nächste zukünftige Ereignis.
        """

        if dataframe.empty:
            return None

        if date_column not in dataframe.columns:
            return None

        jetzt = datetime.now()
        result = dataframe.copy()

        result["_DASHBOARD_DATETIME"] = result.apply(
            lambda row: self._combine_datetime(
                row.get(date_column),
                row.get(time_column)
            ),
            axis=1
        )

        result = result[
            result["_DASHBOARD_DATETIME"].notna()
        ]

        result = result[
            result["_DASHBOARD_DATETIME"] >= jetzt
        ]

        if result.empty:
            return None

        result = result.sort_values(
            "_DASHBOARD_DATETIME"
        )

        record = result.iloc[0].to_dict()
        event_datetime = record.pop(
            "_DASHBOARD_DATETIME",
            None
        )

        record = self._clean_record(record)

        if event_datetime is not None:
            record["_datetime"] = event_datetime

        return record

    def _count_active(self, dataframe):
        """
        Zählt aktive Datensätze.

        Unterstützte Werte:
        JA, AKTIV, TRUE, 1
        """

        if dataframe.empty:
            return 0

        if "AKTIV" not in dataframe.columns:
            return len(dataframe)

        aktive_werte = {
            "JA",
            "AKTIV",
            "TRUE",
            "1",
            "YES"
        }

        return int(
            dataframe["AKTIV"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .isin(aktive_werte)
            .sum()
        )

    def _create_alerts(
        self,
        spiele,
        trainings,
        trainingsplaene,
        mitglieder,
        mannschaften
    ):
        """
        Erstellt erste Dashboard-Hinweise.

        Später ergänzen wir hier echte Konfliktprüfungen,
        Finanzen und fehlende Zuordnungen.
        """

        alerts = []

        if spiele.empty:
            alerts.append({
                "level": "info",
                "area": "Spiele",
                "text": "Keine Spieldaten vorhanden."
            })

        if trainings.empty:
            alerts.append({
                "level": "info",
                "area": "Training",
                "text": "Keine konkreten Trainings vorhanden."
            })

        if trainingsplaene.empty:
            alerts.append({
                "level": "warning",
                "area": "Trainingsplanung",
                "text": "Keine Trainingspläne vorhanden."
            })

        if mitglieder.empty:
            alerts.append({
                "level": "warning",
                "area": "Mitglieder",
                "text": "Keine Mitgliederdaten vorhanden."
            })

        if mannschaften.empty:
            alerts.append({
                "level": "warning",
                "area": "Mannschaften",
                "text": "Keine Mannschaftsdaten vorhanden."
            })

        return alerts

    def _status_level(self, alerts):
        """
        Ermittelt die Ampelfarbe des Dashboards.
        """

        if any(
            alert.get("level") == "critical"
            for alert in alerts
        ):
            return "red"

        if any(
            alert.get("level") == "warning"
            for alert in alerts
        ):
            return "yellow"

        return "green"

    def _records(self, dataframe):
        """
        Wandelt einen DataFrame in bereinigte Datensätze um.
        """

        if dataframe.empty:
            return []

        return [
            self._clean_record(record)
            for record in dataframe.to_dict(
                orient="records"
            )
        ]

    def _clean_record(self, record):
        """
        Entfernt pandas-NaN-Werte aus einem Datensatz.
        """

        cleaned = {}

        for key, value in record.items():

            try:
                if pd.isna(value):
                    cleaned[key] = ""
                    continue
            except (TypeError, ValueError):
                pass

            cleaned[key] = value

        return cleaned

    def _to_date(self, value):
        """
        Wandelt unterschiedliche Excel-Datumswerte in date um.
        """

        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        text = str(value).strip()

        if not text or text.lower() == "nan":
            return None

        for format_string in (
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d/%m/%Y"
        ):
            try:
                return datetime.strptime(
                    text,
                    format_string
                ).date()
            except ValueError:
                continue

        try:
            parsed = pd.to_datetime(
                text,
                dayfirst=True,
                errors="coerce"
            )

            if pd.isna(parsed):
                return None

            return parsed.date()

        except Exception:
            return None

    def _combine_datetime(
        self,
        date_value,
        time_value
    ):
        """
        Verbindet Excel-Datum und Excel-Uhrzeit.
        """

        event_date = self._to_date(date_value)

        if event_date is None:
            return None

        event_time = self._to_time(time_value)

        if event_time is None:
            event_time = datetime.min.time()

        return datetime.combine(
            event_date,
            event_time
        )

    def _to_time(self, value):
        """
        Wandelt unterschiedliche Excel-Zeitwerte in time um.
        """

        if value is None:
            return None

        if isinstance(value, datetime):
            return value.time()

        if hasattr(value, "hour") and hasattr(value, "minute"):
            return value

        text = str(value).strip()

        if not text or text.lower() == "nan":
            return None

        for format_string in (
            "%H:%M:%S",
            "%H:%M"
        ):
            try:
                return datetime.strptime(
                    text,
                    format_string
                ).time()
            except ValueError:
                continue

        return None

    def get_database_stats(self):

        teams = TeamService().get_active()
        facilities = FacilityService().get_active()
        places = PlaceService().get_active()

        return {
            "teams": len(teams),
            "facilities": len(facilities),
            "places": len(places)
        }    