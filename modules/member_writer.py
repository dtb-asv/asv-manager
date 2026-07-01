import shutil
from datetime import datetime
from openpyxl import load_workbook
from modules.id_generator import IdGenerator

from modules.constants import SHEET_MEMBERS, COL_MEMBER_ID


class MemberWriter:

    def _backup(self, excel_datei):
        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_member_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        shutil.copy2(excel_datei, backup)
        return backup

    def _headers(self, ws):
        return {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value
        }

    def _find_row_by_member_id(self, ws, member_id):
        headers = self._headers(ws)
        id_col = headers.get(COL_MEMBER_ID)

        if not id_col:
            return None

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=id_col).value == member_id:
                return row

        return None

    def add_member(self, excel_datei, daten):
        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_MEMBERS]

        headers = self._headers(ws)

        member_id = IdGenerator.next_id(
            excel_datei,
            SHEET_MEMBERS,
            COL_MEMBER_ID,
            "MEMBER"
        )

        daten[COL_MEMBER_ID] = member_id
        neue_zeile = ws.max_row + 1

        for feld, wert in daten.items():
            if feld in headers:
                ws.cell(row=neue_zeile, column=headers[feld]).value = wert

        wb.save(excel_datei)
        return member_id

    def update_member(self, excel_datei, member_id, daten):
        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_MEMBERS]

        headers = self._headers(ws)
        zeile = self._find_row_by_member_id(ws, member_id)

        if zeile is None:
            raise ValueError(f"Mitglied nicht gefunden: {member_id}")

        for feld, wert in daten.items():
            if feld in headers:
                ws.cell(row=zeile, column=headers[feld]).value = wert

        wb.save(excel_datei)

    def archive_member(self, excel_datei, member_id):
        self.update_member(
            excel_datei,
            member_id,
            {
                "STATUS": "Archiviert"
            }
        )