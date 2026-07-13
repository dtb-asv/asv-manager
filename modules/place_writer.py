import shutil
from datetime import datetime
from openpyxl import load_workbook
from modules.constants import SHEET_PLACES


class PlaceWriter:

    def _backup(self, excel_datei):

        backup = excel_datei.replace(
            ".xlsx",
            f"_backup_place_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        shutil.copy2(excel_datei, backup)

        return backup

    def _headers(self, ws):

        return {
            ws.cell(row=1, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value
        }

    def _find_row_by_place_id(self, ws, place_id):

        headers = self._headers(ws)
        id_col = headers.get("PLACE_ID")

        if not id_col:
            return None

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=id_col).value == place_id:
                return row

        return None

    def add_place(self, excel_datei, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_PLACES]

        headers = self._headers(ws)
        neue_zeile = ws.max_row + 1

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=neue_zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

        return daten["PLACE_ID"]

    def update_place(self, excel_datei, place_id, daten):

        self._backup(excel_datei)

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_PLACES]

        headers = self._headers(ws)
        zeile = self._find_row_by_place_id(ws, place_id)

        if zeile is None:
            raise ValueError(f"Platz {place_id} nicht gefunden.")

        for feld, wert in daten.items():

            if feld in headers:
                ws.cell(
                    row=zeile,
                    column=headers[feld]
                ).value = wert

        wb.save(excel_datei)

    def archive_place(self, excel_datei, place_id):

        self.update_place(
            excel_datei,
            place_id,
            {
                "AKTIV": "Nein"
            }
        )