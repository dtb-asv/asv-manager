from openpyxl import load_workbook

from modules.constants import SHEET_MEMBER_ROLES


class MemberRoleWriter:

    def deactivate_roles(self, excel_datei, member_id):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_MEMBER_ROLES]

        headers = [cell.value for cell in ws[1]]

        member_col = headers.index("MEMBER_ID") + 1
        aktiv_col = headers.index("AKTIV") + 1

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=member_col).value == member_id:
                ws.cell(row=row, column=aktiv_col).value = "NEIN"

        wb.save(excel_datei)

    def add_role(self, excel_datei, member_id, role_code):

        wb = load_workbook(excel_datei)
        ws = wb[SHEET_MEMBER_ROLES]

        ws.append([
            member_id,
            role_code,
            "",
            "",
            "JA",
            ""
        ])

        wb.save(excel_datei)