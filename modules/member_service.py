import pandas as pd

from datetime import datetime
from openpyxl import load_workbook
from modules.constants import SHEET_MEMBERS
from modules.member_writer import MemberWriter
from modules.history_service import HistoryService
from modules.id_generator import IdGenerator
from modules.constants import COL_MEMBER_ID
from modules.repositories.member_repository import MemberRepository


class MemberService:

    def __init__(self):
        self.writer = MemberWriter()
        self.history = HistoryService()
        self.repo = MemberRepository()

    def get_all(self):
        return self.repo.get_all()  

    def get_by_id(self, person_id):
        return self.repo.get_by_id(person_id)    

    def create_member(self, member):
        return self.repo.save(member)


    def update_member(self, person_id, member):
        member["person_id"] = person_id
        return self.repo.save(member)   

    def archive_member(self, person_id):
        self.repo.archive(person_id)       

    def load_members(self, excel_datei):
        wb = load_workbook(excel_datei)

        if SHEET_MEMBERS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_MEMBERS)
            ws.append([
                "MEMBER_ID",
                "VORNAME",
                "NACHNAME",
                "GEBURTSDATUM",
                "EINTRITT",
                "AUSTRITT",
                "STATUS",
                "BEMERKUNG",
                "GESCHLECHT",
                "MOBIL",
                "SPIELERPASSNUMMER",
                "EMAIL"
            ])
            wb.save(excel_datei)

        return pd.read_excel(
            excel_datei,
            sheet_name=SHEET_MEMBERS
        )

    def next_member_id(self, excel_datei):
        df = self.load_members(excel_datei)

        if df.empty or "MEMBER_ID" not in df.columns:
            return "MEMBER000001"

        nummern = []

        for value in df["MEMBER_ID"].dropna():
            text = str(value)
            if text.startswith("MEMBER"):
                nummern.append(int(text.replace("MEMBER", "")))

        next_number = max(nummern, default=0) + 1
        return f"MEMBER{next_number:06d}"

    def add_member(self, excel_datei, daten):

        daten[COL_MEMBER_ID] = IdGenerator.next_id(
            excel_datei,
            SHEET_MEMBERS,
            COL_MEMBER_ID,
            "MEMBER"
        )

        daten["EINTRITT"] = datetime.now().strftime("%d.%m.%Y")
        daten["AUSTRITT"] = "31.12.9999"
        daten["STATUS"] = "Aktiv"

        return self.writer.add_member(
            excel_datei,
            daten
        )