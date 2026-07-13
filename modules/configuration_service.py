from openpyxl import load_workbook
from modules.constants import (
    SHEET_CFG_ROLES,
    SHEET_CFG_RIGHTS,
    SHEET_CFG_ROLE_RIGHTS,
    SHEET_CFG_LOOKUPS,
    SHEET_MEMBER_ROLES,
    SHEET_TRAININGS,
    SHEET_TEAM_ASSIGNMENTS,
    SHEET_TRAINING_PARTICIPANTS,
    SHEET_TRAINING_SCHEDULES,
    SHEET_CLUB,
    SHEET_SEASONS,
    SHEET_FACILITIES,
    SHEET_PLACES,
    SHEET_TEAMS,
    SHEET_TEAM_MEMBERS,
    SHEET_DEPARTMENTS
)


class ConfigurationService:

    def ensure_configuration_sheets(self, excel_datei):

        wb = load_workbook(excel_datei)

        if SHEET_CFG_ROLES not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CFG_ROLES)
            ws.append([
                "ROLE_ID",
                "NAME",
                "BESCHREIBUNG",
                "AKTIV"
            ])

        if SHEET_CFG_RIGHTS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CFG_RIGHTS)
            ws.append([
                "RIGHT_ID",
                "NAME",
                "BESCHREIBUNG",
                "AKTIV"
            ])

        if SHEET_CFG_ROLE_RIGHTS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CFG_ROLE_RIGHTS)
            ws.append([
                "ROLE_ID",
                "RIGHT_ID"
            ])

        self.ensure_default_roles(wb)   

        if SHEET_CFG_LOOKUPS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CFG_LOOKUPS)
            ws.append([
                "LOOKUP_ID",
                "LOOKUP_TYPE",
                "CODE",
                "NAME",
                "BESCHREIBUNG",
                "SORTIERUNG",
                "AKTIV",
                "SYSTEM"
            ]) 

        self.ensure_default_lookups(wb)    

        if SHEET_MEMBER_ROLES not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_MEMBER_ROLES)
            ws.append([
                "MEMBER_ID",
                "ROLE_CODE",
                "VON",
                "BIS",
                "AKTIV",
                "BEMERKUNG"
            ])

        if SHEET_TEAM_ASSIGNMENTS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_TEAM_ASSIGNMENTS)
            ws.append([
                "TEAM_ID",
                "MEMBER_ID",
                "ROLE_CODE",
                "VON",
                "BIS",
                "AKTIV",
                "BEMERKUNG"
            ])  

        if SHEET_TRAININGS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_TRAININGS)
            ws.append([
                "TRAINING_ID",
                "TEAM_ID",
                "DATUM",
                "STARTZEIT",
                "ENDZEIT",
                "ORT",
                "TRAINING_TYPE",
                "STATUS",
                "AKTIV",
                "BEMERKUNG"
            ])      

        if SHEET_TRAINING_PARTICIPANTS not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_TRAINING_PARTICIPANTS)

            ws.append([
                "TRAINING_ID",
                "MEMBER_ID",
                "ROLE",
                "STATUS",
                "BEMERKUNG",
                "AKTIV"
            ]) 

        if SHEET_TRAINING_SCHEDULES not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_TRAINING_SCHEDULES)

            ws.append([
                "SCHEDULE_ID",
                "TEAM_ID",
                "SAISON",
                "WOCHENTAG",
                "BEGINN",
                "ENDE",
                "PLATZ",
                "TRAINING_TYPE",
                "AKTIV",
                "BEMERKUNG"
            ])    

        if SHEET_CLUB not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CLUB)
            ws.append([
                "CLUB_ID",
                "NAME",
                "LANDESVERBAND",
                "LOGO",
                "AKTIV"
            ])

        if SHEET_SEASONS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_SEASONS)
            ws.append([
                "SAISON_ID",
                "NAME",
                "VON",
                "BIS",
                "AKTIV",
                "BEMERKUNG"
            ])

        if SHEET_FACILITIES not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_FACILITIES)
            ws.append([
                "FACILITY_ID",
                "NAME",
                "ORT",
                "ADRESSE",
                "AKTIV",
                "BEMERKUNG"
            ])

        if SHEET_PLACES not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_PLACES)

            ws.append([
                "PLACE_ID",
                "FACILITY_ID",
                "NAME",
                "TRAINING_KAPAZITAET",
                "IST_TRAININGSPLATZ",
                "IST_SPIELPLATZ",
                "SPIEL_KAPAZITAET",
                "REIHENFOLGE",
                "AKTIV",
                "BEMERKUNG"
            ])

        if SHEET_TEAMS not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_TEAMS)

            ws.append([
                "TEAM_ID",
                "NAME",
                "ALTERSKLASSE",
                "DEPARTMENT_ID",
                "AKTIV",
                "BEMERKUNG"
            ])    

        if SHEET_TEAM_MEMBERS not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_TEAM_MEMBERS)

            ws.append([
                "TEAM_MEMBER_ID",
                "TEAM_ID",
                "MEMBER_ID",
                "ROLLE",
                "VON",
                "BIS",
                "AKTIV",
                "BEMERKUNG"
            ])     

        if SHEET_DEPARTMENTS not in wb.sheetnames:

            ws = wb.create_sheet(SHEET_DEPARTMENTS)

            ws.append([
                "DEPARTMENT_ID",
                "NAME",
                "LEITER_MEMBER_ID",
                "AKTIV",
                "BEMERKUNG"
            ])      

        wb.save(excel_datei)

    def ensure_default_lookups(self, wb):

        ws = wb[SHEET_CFG_LOOKUPS]

        if ws.max_row > 1:
            return

        lookups = [
            ("LOOKUP000001", "ROLE", "SPIELER", "Spieler", "Aktiver Spieler", 10, "JA", "JA"),
            ("LOOKUP000002", "ROLE", "TRAINER", "Trainer", "Trainer", 20, "JA", "JA"),
            ("LOOKUP000003", "ROLE", "CO_TRAINER", "Co-Trainer", "Co-Trainer", 30, "JA", "JA"),
            ("LOOKUP000004", "ROLE", "TORMANNTRAINER", "Tormanntrainer", "Tormanntrainer", 40, "JA", "JA"),
            ("LOOKUP000005", "ROLE", "NACHWUCHSLEITER", "Nachwuchsleiter", "Nachwuchsleiter", 50, "JA", "JA"),

            ("LOOKUP000006", "STATUS", "AKTIV", "Aktiv", "Aktiv", 10, "JA", "JA"),
            ("LOOKUP000007", "STATUS", "ARCHIVIERT", "Archiviert", "Archiviert", 20, "JA", "JA"),

            ("LOOKUP000008", "GAME_TYPE", "SPIEL", "Spiel", "Spiel", 10, "JA", "JA"),
            ("LOOKUP000009", "GAME_TYPE", "FREUNDSCHAFTSSPIEL", "Freundschaftsspiel", "Freundschaftsspiel", 20, "JA", "JA"),
            ("LOOKUP000010", "GAME_TYPE", "CAMP", "Camp", "Camp", 30, "JA", "JA"),
        ]

        for item in lookups:
            ws.append(list(item))    

    def ensure_default_roles(self, wb):

        ws = wb[SHEET_CFG_ROLES]

        if ws.max_row > 1:
            return

        rollen = [
            ("ROLE000001", "Spieler", "Aktiver Spieler"),
            ("ROLE000002", "Trainer", "Haupttrainer"),
            ("ROLE000003", "Co-Trainer", "Co-Trainer"),
            ("ROLE000004", "Tormanntrainer", "Torwarttrainer"),
            ("ROLE000005", "Betreuer", "Betreuer"),
            ("ROLE000006", "Nachwuchsleiter", "Nachwuchsleiter"),
            ("ROLE000007", "Jugendleiter", "Jugendleiter"),
            ("ROLE000008", "Sportlicher Leiter", "Sportlicher Leiter"),
            ("ROLE000009", "Kassier", "Kassier"),
            ("ROLE000010", "Obmann", "Obmann"),
            ("ROLE000011", "Funktionär", "Funktionär"),
            ("ROLE000012", "Turnierleitung", "Turnierleitung"),
            ("ROLE000013", "Social Media", "Social Media"),
            ("ROLE000014", "Platzwart", "Platzwart")
        ]

        for role_id, name, beschreibung in rollen:

            ws.append([
                role_id,
                name,
                beschreibung,
                "JA"
            ])    